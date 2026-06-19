"""Excel export utilities for detection history."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .history_store import DetectionRecord
from .paths import EXPORT_OUTPUT_DIR, ensure_desktop_dirs


def export_records_to_excel(records: list[DetectionRecord], output_path: Path | None = None) -> Path:
    """Export detection records to an xlsx workbook."""
    ensure_desktop_dirs()
    path = output_path or EXPORT_OUTPUT_DIR / f"history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Detection History"

    headers = [
        "ID",
        "时间",
        "模式",
        "输入源",
        "输出文件",
        "模型",
        "设备",
        "性能",
        "总数量",
        "类别统计",
        "状态",
    ]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for record in records:
        classes = ", ".join(f"{name}: {count}" for name, count in record.class_counts.items()) or "-"
        sheet.append(
            [
                record.id,
                record.created_at,
                record.mode,
                record.source_path,
                record.output_path,
                record.model_path,
                record.device,
                _format_performance(record),
                record.total_count,
                classes,
                record.status,
            ]
        )

    widths = [8, 20, 14, 36, 36, 36, 18, 10, 10, 32, 12]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A2"
    workbook.save(path)
    return path


def _format_performance(record: DetectionRecord) -> str:
    if record.performance_unit.lower() == "ms":
        return f"{record.fps:.0f} ms"
    return f"{record.fps:.1f} FPS"

