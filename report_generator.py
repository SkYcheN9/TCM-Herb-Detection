"""Generate experiment materials and a Word report from YOLO training outputs.

The generator reads the project training artifacts under runs/ and ablation
report directories, then exports CSV, Excel, PNG figures, confusion matrices,
and a DOCX report suitable for experiment documentation.
"""

from __future__ import annotations

import argparse
import csv
import math
import shutil
import sys
import zipfile
from dataclasses import dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape

ROOT = Path(__file__).resolve().parent


LOSS_METRICS = [
    ("box_loss", "train/box_loss", "val/box_loss", "Box Loss"),
    ("cls_loss", "train/cls_loss", "val/cls_loss", "Classification Loss"),
    ("dfl_loss", "train/dfl_loss", "val/dfl_loss", "DFL Loss"),
]

MAP_METRICS = [
    ("map50", "metrics/mAP50(B)", "mAP50"),
    ("map50_95", "metrics/mAP50-95(B)", "mAP50-95"),
]

ABLATION_FIELDS = [
    "Rank",
    "Experiment",
    "Config",
    "Suite",
    "CBAM",
    "BiFPN",
    "FocalLoss",
    "GhostConv",
    "DecoupledHead",
    "Init",
    "Epochs",
    "ImageSize",
    "Batch",
    "Device",
    "Precision",
    "Recall",
    "mAP50",
    "mAP50-95",
    "FPS",
    "Delta_mAP50",
    "Delta_mAP50-95",
    "Delta_FPS",
    "Weights",
]


@dataclass
class FigureSet:
    """Generated figure paths grouped by report section."""

    loss: dict[str, Path] = field(default_factory=dict)
    map: dict[str, Path] = field(default_factory=dict)
    ablation: dict[str, Path] = field(default_factory=dict)
    confusion: dict[str, Path] = field(default_factory=dict)


@dataclass
class SourceInfo:
    """Resolved project artifact locations."""

    run_dir: Path
    results_csv: Path
    args_yaml: Path | None
    ablation_dir: Path | None
    ablation_summary: Path | None
    ablation_history: Path | None


def parse_args() -> argparse.Namespace:
    """Parse command-line options."""

    parser = argparse.ArgumentParser(
        description="Generate YOLO experiment CSV, Excel, PNG, and DOCX reports."
    )
    parser.add_argument("--runs-dir", default="runs", help="Directory containing YOLO runs.")
    parser.add_argument("--run-dir", default=None, help="Specific YOLO run directory to report.")
    parser.add_argument(
        "--ablation-dir",
        default=None,
        help="Specific ablation output directory. If omitted, the most complete one is selected.",
    )
    parser.add_argument(
        "--output",
        default="reports/generated_experiment_materials",
        help="Output directory for generated materials.",
    )
    parser.add_argument(
        "--project-name",
        default="TCM-SliceAI 中医药饮片智能检测与识别系统",
        help="Project name used in the Word report.",
    )
    parser.add_argument(
        "--report-title",
        default="基于改进 YOLOv8 的中医药饮片检测实验报告",
        help="Word report title.",
    )
    return parser.parse_args()


def resolve_path(value: str | Path | None) -> Path | None:
    """Resolve a path relative to the repository root."""

    if value is None:
        return None
    path = Path(value)
    if path.is_absolute():
        return path
    return ROOT / path


def read_csv_rows(path: Path | None) -> list[dict[str, str]]:
    """Read a CSV file into dictionaries."""

    if path is None or not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        return list(csv.DictReader(file))


def write_csv_rows(path: Path, rows: list[dict[str, Any]], fields: list[str] | None = None) -> None:
    """Write dictionaries to CSV."""

    path.parent.mkdir(parents=True, exist_ok=True)
    if fields is None:
        fields = []
        for row in rows:
            for key in row:
                if key not in fields:
                    fields.append(key)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def numeric(value: Any) -> float | None:
    """Parse a float value from a CSV cell."""

    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        parsed = float(text)
    except ValueError:
        return None
    if math.isnan(parsed) or math.isinf(parsed):
        return None
    return parsed


def fmt(value: Any, digits: int = 4) -> str:
    """Format numbers for report text."""

    parsed = numeric(value)
    if parsed is None:
        return "-" if value in (None, "") else str(value)
    return f"{parsed:.{digits}f}"


def fmt_delta(value: Any, digits: int = 4) -> str:
    """Format a signed delta."""

    parsed = numeric(value)
    if parsed is None:
        return "-"
    return f"{parsed:+.{digits}f}"


def truthy(value: Any) -> bool:
    """Return whether a CSV boolean-like value is true."""

    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


def read_simple_yaml(path: Path | None) -> dict[str, Any]:
    """Read a small YAML file, using PyYAML when available."""

    if path is None or not path.exists():
        return {}
    try:
        import yaml  # type: ignore

        loaded = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
        return loaded if isinstance(loaded, dict) else {}
    except Exception:
        data: dict[str, Any] = {}
        for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
            if ":" not in line or line.lstrip().startswith("#"):
                continue
            key, value = line.split(":", 1)
            data[key.strip()] = value.strip().strip("'\"")
        return data


def find_results_csv(run_dir: Path) -> Path | None:
    """Return the results.csv path for a YOLO run directory."""

    direct = run_dir / "results.csv"
    if direct.exists():
        return direct
    candidates = sorted(run_dir.rglob("results.csv"), key=lambda item: item.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


def choose_run_dir(runs_dir: Path, explicit: Path | None) -> Path:
    """Choose the primary run directory used for training curves."""

    if explicit is not None:
        results = find_results_csv(explicit)
        if results is None:
            raise FileNotFoundError(f"No results.csv found under {explicit}")
        return explicit
    baseline = runs_dir / "baseline"
    if (baseline / "results.csv").exists():
        return baseline
    candidates = [path.parent for path in runs_dir.rglob("results.csv")] if runs_dir.exists() else []
    if not candidates:
        raise FileNotFoundError(f"No results.csv found under {runs_dir}")
    return sorted(candidates, key=lambda item: item.stat().st_mtime, reverse=True)[0]


def count_csv_rows(path: Path | None) -> int:
    """Count rows in a CSV file."""

    return len(read_csv_rows(path))


def ablation_candidates(explicit: Path | None) -> list[Path]:
    """Return candidate ablation directories in priority-search order."""

    if explicit is not None:
        return [explicit]
    candidates = [
        ROOT / "ablation",
        ROOT / "reports" / "ablation",
        ROOT / "final_results_full" / "reports" / "ablation",
    ]
    candidates.extend(sorted((ROOT / "reports").glob("*ablation*")) if (ROOT / "reports").exists() else [])
    candidates.extend(
        sorted((ROOT / "final_results_full" / "reports").glob("*ablation*"))
        if (ROOT / "final_results_full" / "reports").exists()
        else []
    )
    unique: list[Path] = []
    for item in candidates:
        if item not in unique:
            unique.append(item)
    return unique


def choose_ablation_dir(explicit: Path | None) -> Path | None:
    """Choose the ablation directory with the richest available result set."""

    scored: list[tuple[int, float, Path]] = []
    for candidate in ablation_candidates(explicit):
        summary = candidate / "summary.csv"
        history = candidate / "history.csv"
        if not summary.exists() and not history.exists():
            continue
        score = count_csv_rows(summary) * 1000 + count_csv_rows(history)
        mtime = max(
            [p.stat().st_mtime for p in [summary, history] if p.exists()],
            default=candidate.stat().st_mtime if candidate.exists() else 0.0,
        )
        scored.append((score, mtime, candidate))
    if not scored:
        return None
    return sorted(scored, reverse=True)[0][2]


def resolve_sources(args: argparse.Namespace) -> SourceInfo:
    """Resolve all input artifact paths."""

    runs_dir = resolve_path(args.runs_dir)
    assert runs_dir is not None
    run_dir = choose_run_dir(runs_dir, resolve_path(args.run_dir))
    results_csv = find_results_csv(run_dir)
    if results_csv is None:
        raise FileNotFoundError(f"No results.csv found under {run_dir}")
    ablation_dir = choose_ablation_dir(resolve_path(args.ablation_dir))
    return SourceInfo(
        run_dir=run_dir,
        results_csv=results_csv,
        args_yaml=run_dir / "args.yaml" if (run_dir / "args.yaml").exists() else None,
        ablation_dir=ablation_dir,
        ablation_summary=(ablation_dir / "summary.csv") if ablation_dir and (ablation_dir / "summary.csv").exists() else None,
        ablation_history=(ablation_dir / "history.csv") if ablation_dir and (ablation_dir / "history.csv").exists() else None,
    )


def row_value(row: dict[str, str], key: str) -> float | None:
    """Return a numeric row value for a metric key."""

    return numeric(row.get(key))


def metric_series(rows: list[dict[str, str]], key: str) -> tuple[list[float], list[float]]:
    """Return epoch and metric series for rows where the metric exists."""

    xs: list[float] = []
    ys: list[float] = []
    for index, row in enumerate(rows, start=1):
        y = row_value(row, key)
        if y is None:
            continue
        x = row_value(row, "epoch")
        xs.append(x if x is not None else float(index))
        ys.append(y)
    return xs, ys


def require_matplotlib() -> Any:
    """Import matplotlib with a non-interactive backend."""

    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt

        return plt
    except Exception as exc:  # pragma: no cover - environment dependent
        raise RuntimeError(
            "matplotlib is required to generate PNG figures. Install it with: pip install matplotlib"
        ) from exc


def style_axis(ax: Any, title: str, ylabel: str) -> None:
    """Apply a consistent chart style."""

    ax.set_title(title, fontsize=14, weight="bold")
    ax.set_xlabel("Epoch")
    ax.set_ylabel(ylabel)
    ax.grid(True, linestyle="--", alpha=0.28)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)


def save_line_chart(
    rows: list[dict[str, str]],
    series: list[tuple[str, str]],
    title: str,
    ylabel: str,
    output: Path,
) -> Path:
    """Save a line chart for one or more metric series."""

    plt = require_matplotlib()
    output.parent.mkdir(parents=True, exist_ok=True)
    fig, ax = plt.subplots(figsize=(9.6, 5.4), dpi=160)
    plotted = False
    for key, label in series:
        xs, ys = metric_series(rows, key)
        if not ys:
            continue
        ax.plot(xs, ys, linewidth=2.0, label=label)
        plotted = True
    style_axis(ax, title, ylabel)
    if plotted:
        ax.legend(frameon=False)
    fig.tight_layout()
    fig.savefig(output)
    plt.close(fig)
    return output


def save_ablation_bar(rows: list[dict[str, Any]], metric: str, title: str, output: Path) -> Path | None:
    """Save an ablation bar chart sorted by a target metric."""

    values = [(row.get("Experiment", ""), numeric(row.get(metric))) for row in rows]
    values = [(str(name), value) for name, value in values if value is not None]
    if not values:
        return None
    values.sort(key=lambda item: item[1], reverse=True)
    plt = require_matplotlib()
    height = max(5.2, 0.45 * len(values) + 1.4)
    fig, ax = plt.subplots(figsize=(10.5, height), dpi=160)
    names = [name for name, _ in values]
    scores = [value for _, value in values]
    bars = ax.barh(range(len(values)), scores, color="#2b7a78")
    ax.set_yticks(range(len(values)))
    ax.set_yticklabels(names, fontsize=9)
    ax.invert_yaxis()
    ax.set_xlabel(metric)
    ax.set_title(title, fontsize=14, weight="bold")
    ax.grid(axis="x", linestyle="--", alpha=0.25)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    for bar, value in zip(bars, scores):
        ax.text(value + 0.002, bar.get_y() + bar.get_height() / 2, fmt(value), va="center", fontsize=8.5)
    fig.tight_layout()
    output.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output)
    plt.close(fig)
    return output


def save_speed_accuracy_scatter(rows: list[dict[str, Any]], output: Path) -> Path | None:
    """Save an FPS versus mAP50-95 scatter plot for ablation rows."""

    points = []
    for row in rows:
        fps = numeric(row.get("FPS"))
        map95 = numeric(row.get("mAP50-95"))
        if fps is not None and map95 is not None:
            points.append((str(row.get("Experiment", "")), fps, map95))
    if not points:
        return None
    plt = require_matplotlib()
    fig, ax = plt.subplots(figsize=(9.6, 5.6), dpi=160)
    xs = [item[1] for item in points]
    ys = [item[2] for item in points]
    ax.scatter(xs, ys, s=58, color="#c44536", alpha=0.9)
    for name, x, y in points:
        ax.annotate(name, (x, y), xytext=(5, 4), textcoords="offset points", fontsize=8)
    ax.set_title("Accuracy-Speed Trade-off", fontsize=14, weight="bold")
    ax.set_xlabel("FPS")
    ax.set_ylabel("mAP50-95")
    ax.grid(True, linestyle="--", alpha=0.25)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout()
    output.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output)
    plt.close(fig)
    return output


def save_multi_experiment_chart(
    rows: list[dict[str, str]],
    metric: str,
    title: str,
    output: Path,
    ylabel: str,
) -> Path | None:
    """Save a line chart grouped by the history.csv experiment column."""

    grouped: dict[str, list[dict[str, str]]] = {}
    for row in rows:
        experiment = row.get("experiment") or row.get("Experiment") or "Experiment"
        grouped.setdefault(experiment, []).append(row)
    if not grouped:
        return None
    plt = require_matplotlib()
    fig, ax = plt.subplots(figsize=(10.5, 5.8), dpi=160)
    plotted = False
    for experiment, group_rows in grouped.items():
        xs, ys = metric_series(group_rows, metric)
        if not ys:
            continue
        ax.plot(xs, ys, linewidth=1.8, label=experiment)
        plotted = True
    if not plotted:
        plt.close(fig)
        return None
    style_axis(ax, title, ylabel)
    ax.legend(frameon=False, fontsize=8, ncols=2)
    fig.tight_layout()
    output.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output)
    plt.close(fig)
    return output


def enrich_ablation_rows(rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    """Add rank and baseline deltas to ablation summary rows."""

    if not rows:
        return []
    baseline = next((row for row in rows if str(row.get("Experiment", "")).lower() == "baseline"), rows[0])
    base_map50 = numeric(baseline.get("mAP50")) or 0.0
    base_map95 = numeric(baseline.get("mAP50-95")) or 0.0
    base_fps = numeric(baseline.get("FPS")) or 0.0
    ranked = sorted(rows, key=lambda row: numeric(row.get("mAP50-95")) or -1.0, reverse=True)
    rank_by_name = {str(row.get("Experiment", "")): index for index, row in enumerate(ranked, start=1)}
    enriched: list[dict[str, Any]] = []
    for row in rows:
        item: dict[str, Any] = dict(row)
        item["Rank"] = rank_by_name.get(str(row.get("Experiment", "")), "")
        item["Delta_mAP50"] = (numeric(row.get("mAP50")) or 0.0) - base_map50
        item["Delta_mAP50-95"] = (numeric(row.get("mAP50-95")) or 0.0) - base_map95
        item["Delta_FPS"] = (numeric(row.get("FPS")) or 0.0) - base_fps
        enriched.append(item)
    return sorted(enriched, key=lambda row: numeric(row.get("Rank")) or 9999)



def infer_experiment_fields(name: str) -> dict[str, str]:
    """Infer ablation metadata from a display name."""

    lower_name = name.lower()
    key = lower_name.replace("baseline+", "").replace("+", "_").replace("-", "_")
    config_map = {
        "baseline": "configs/baseline.yaml",
        "cbam": "configs/cbam.yaml",
        "bifpn": "configs/bifpn.yaml",
        "focal": "configs/focal.yaml",
        "ghostconv": "configs/ghostconv.yaml",
        "decoupledhead": "configs/decoupled_head.yaml",
        "cbam_bifpn": "configs/cbam_bifpn.yaml",
        "cbam_bifpn_focal": "configs/cbam_bifpn_focal.yaml",
        "cbam_bifpn_ghostconv": "configs/cbam_bifpn_ghost.yaml",
        "cbam_bifpn_ghostconv_decoupledhead": "configs/cbam_bifpn_ghost_decoupled.yaml",
        "fullmodel": "configs/full_model.yaml",
    }
    has_cbam_bifpn = "cbam+bifpn" in lower_name
    return {
        "Experiment": name,
        "Config": config_map.get(key, ""),
        "Suite": "candidate" if "ghostconv" in lower_name and has_cbam_bifpn else "final_selection",
        "CBAM": str("cbam" in lower_name or "fullmodel" in lower_name),
        "BiFPN": str("bifpn" in lower_name or "fullmodel" in lower_name),
        "FocalLoss": str("focal" in lower_name or "fullmodel" in lower_name),
        "GhostConv": str("ghostconv" in lower_name or "fullmodel" in lower_name),
        "DecoupledHead": str("decoupledhead" in lower_name or "fullmodel" in lower_name),
        "Init": "pretrained",
    }


def read_final_selection_rows(path: Path | None = None) -> list[dict[str, str]]:
    """Read optional final selection table rows from project documentation."""

    selection_path = path or (ROOT / "docs" / "FINAL_MODEL_SELECTION.md")
    if not selection_path.exists():
        return []
    rows: list[dict[str, str]] = []
    for line in selection_path.read_text(encoding="utf-8", errors="ignore").splitlines():
        if not line.startswith("|") or "---" in line:
            continue
        parts = [part.strip() for part in line.strip().strip("|").split("|")]
        if len(parts) < 7 or not parts[0].isdigit():
            continue
        name = parts[1]
        row = infer_experiment_fields(name)
        row.update(
            {
                "Precision": parts[4],
                "Recall": parts[5],
                "mAP50": parts[2],
                "mAP50-95": parts[3],
                "FPS": parts[6],
            }
        )
        rows.append(row)
    return rows


def merge_final_selection_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    """Append final-selection rows that are absent from summary.csv."""

    merged = list(rows)
    known = {str(row.get("Experiment", "")) for row in merged}
    for row in read_final_selection_rows():
        if row.get("Experiment") not in known:
            merged.append(row)
            known.add(str(row.get("Experiment", "")))
    return merged
def best_row(rows: list[dict[str, Any]], metric: str) -> dict[str, Any] | None:
    """Return the best row for a metric."""

    valid = [row for row in rows if numeric(row.get(metric)) is not None]
    if not valid:
        return None
    return max(valid, key=lambda row: numeric(row.get(metric)) or -1.0)


def best_epoch(rows: list[dict[str, str]]) -> dict[str, Any]:
    """Summarize the best and final epoch for the primary training run."""

    best = best_row(rows, "metrics/mAP50-95(B)")
    final = rows[-1] if rows else {}
    return {
        "best_epoch": best.get("epoch") if best else "",
        "best_mAP50": best.get("metrics/mAP50(B)") if best else "",
        "best_mAP50-95": best.get("metrics/mAP50-95(B)") if best else "",
        "final_epoch": final.get("epoch", ""),
        "final_box_loss": final.get("train/box_loss", ""),
        "final_cls_loss": final.get("train/cls_loss", ""),
        "final_dfl_loss": final.get("train/dfl_loss", ""),
        "final_mAP50": final.get("metrics/mAP50(B)", ""),
        "final_mAP50-95": final.get("metrics/mAP50-95(B)", ""),
    }


def copy_confusion_matrices(source: SourceInfo, figures_dir: Path) -> dict[str, Path]:
    """Copy confusion matrices from the primary run into the generated figures directory."""

    copied: dict[str, Path] = {}
    for name in ["confusion_matrix.png", "confusion_matrix_normalized.png"]:
        src = source.run_dir / name
        if not src.exists() and source.ablation_dir:
            candidates = sorted(source.ablation_dir.rglob(name), key=lambda item: item.stat().st_mtime, reverse=True)
            src = candidates[0] if candidates else src
        if src.exists():
            dst = figures_dir / name
            dst.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(src, dst)
            copied[name] = dst
    return copied


def generate_figures(
    training_rows: list[dict[str, str]],
    history_rows: list[dict[str, str]],
    ablation_rows: list[dict[str, Any]],
    source: SourceInfo,
    figures_dir: Path,
) -> FigureSet:
    """Generate all requested PNG figures."""

    figures = FigureSet()
    for file_key, train_key, val_key, title in LOSS_METRICS:
        series = [(train_key, f"train/{file_key}")]
        if any(row.get(val_key) not in (None, "") for row in training_rows):
            series.append((val_key, f"val/{file_key}"))
        figures.loss[file_key] = save_line_chart(
            training_rows,
            series,
            title=f"{title} Curve",
            ylabel=title,
            output=figures_dir / f"{file_key}_curve.png",
        )
    figures.loss["loss_all"] = save_line_chart(
        training_rows,
        [(train_key, label) for label, train_key, _, _ in LOSS_METRICS],
        title="Training Loss Curves",
        ylabel="Loss",
        output=figures_dir / "loss_curve.png",
    )
    for file_key, metric_key, title in MAP_METRICS:
        figures.map[file_key] = save_line_chart(
            training_rows,
            [(metric_key, title)],
            title=f"{title} Curve",
            ylabel=title,
            output=figures_dir / f"{file_key}_curve.png",
        )
    figures.map["map_all"] = save_line_chart(
        training_rows,
        [(metric_key, title) for _, metric_key, title in MAP_METRICS],
        title="mAP Curves",
        ylabel="mAP",
        output=figures_dir / "map_curve.png",
    )
    if history_rows:
        loss_history = save_multi_experiment_chart(
            history_rows,
            "train/box_loss",
            "Ablation Box Loss Curves",
            figures_dir / "ablation_box_loss_curve.png",
            "Box Loss",
        )
        map_history = save_multi_experiment_chart(
            history_rows,
            "metrics/mAP50-95(B)",
            "Ablation mAP50-95 Curves",
            figures_dir / "ablation_map50_95_curve.png",
            "mAP50-95",
        )
        if loss_history:
            figures.ablation["ablation_box_loss"] = loss_history
        if map_history:
            figures.ablation["ablation_map50_95"] = map_history
    bar = save_ablation_bar(
        ablation_rows,
        "mAP50-95",
        "Ablation Ranking by mAP50-95",
        figures_dir / "ablation_map50_95_bar.png",
    )
    if bar:
        figures.ablation["ablation_map50_95_bar"] = bar
    scatter = save_speed_accuracy_scatter(ablation_rows, figures_dir / "ablation_speed_accuracy.png")
    if scatter:
        figures.ablation["ablation_speed_accuracy"] = scatter
    figures.confusion = copy_confusion_matrices(source, figures_dir)
    return figures


def dataset_summary() -> dict[str, Any]:
    """Collect lightweight dataset information for the report."""

    data_yaml = ROOT / "dataset_augmented" / "data.yaml"
    if not data_yaml.exists():
        data_yaml = ROOT / "dataset" / "data.yaml"
    data = read_simple_yaml(data_yaml)
    names = data.get("names", [])
    if isinstance(names, dict):
        names = [names[key] for key in sorted(names)]
    if not isinstance(names, list):
        names = []
    counts = {
        "dataset_yaml": str(data_yaml.relative_to(ROOT)) if data_yaml.exists() else "",
        "classes": len(names),
        "class_names": ", ".join(str(name) for name in names),
        "train_images": count_images(ROOT / "dataset_augmented" / "images" / "train"),
        "val_images": count_images(ROOT / "dataset_augmented" / "images" / "val"),
    }
    if counts["train_images"] == 0:
        counts["train_images"] = count_images(ROOT / "dataset" / "images" / "train")
    if counts["val_images"] == 0:
        counts["val_images"] = count_images(ROOT / "dataset" / "images" / "val")
    counts["total_images"] = counts["train_images"] + counts["val_images"]
    return counts


def count_images(directory: Path) -> int:
    """Count common image files under a directory."""

    if not directory.exists():
        return 0
    suffixes = {".jpg", ".jpeg", ".png", ".bmp", ".webp"}
    return sum(1 for item in directory.rglob("*") if item.suffix.lower() in suffixes)


def add_excel_sheet(wb: Any, title: str, rows: list[dict[str, Any]], fields: list[str] | None = None) -> None:
    """Add a worksheet from row dictionaries."""

    ws = wb.create_sheet(title[:31])
    if fields is None:
        fields = []
        for row in rows:
            for key in row:
                if key not in fields:
                    fields.append(key)
    if not fields:
        ws.append(["No data"])
        return
    ws.append(fields)
    for row in rows:
        ws.append([row.get(field, "") for field in fields])
    ws.freeze_panes = "A2"
    for column_cells in ws.columns:
        max_len = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
        ws.column_dimensions[column_letter].width = min(max(max_len + 2, 10), 48)


def export_excel(
    output: Path,
    training_rows: list[dict[str, str]],
    ablation_rows: list[dict[str, Any]],
    history_rows: list[dict[str, str]],
    source_rows: list[dict[str, str]],
    figure_paths: list[Path],
) -> None:
    """Export a multi-sheet Excel workbook."""

    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as ExcelImage
        from openpyxl.styles import Font, PatternFill
    except Exception as exc:  # pragma: no cover - environment dependent
        raise RuntimeError("openpyxl is required to generate Excel files. Install it with: pip install openpyxl") from exc

    wb = Workbook()
    default = wb.active
    wb.remove(default)
    add_excel_sheet(wb, "Training Metrics", training_rows)
    add_excel_sheet(wb, "Ablation Summary", ablation_rows, ABLATION_FIELDS)
    add_excel_sheet(wb, "Ablation History", history_rows)
    add_excel_sheet(wb, "Sources", source_rows)
    for ws in wb.worksheets:
        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2B7A78")
            ws.auto_filter.ref = ws.dimensions
    figures_ws = wb.create_sheet("Figures")
    figures_ws.append(["Figure", "Path"])
    for index, fig_path in enumerate(figure_paths, start=2):
        figures_ws.cell(row=index, column=1).value = fig_path.stem
        figures_ws.cell(row=index, column=2).value = str(fig_path)
    anchor_row = 2
    for fig_path in figure_paths[:6]:
        try:
            image = ExcelImage(str(fig_path))
            image.width = min(image.width, 720)
            image.height = int(image.height * (image.width / max(image.width, 1)))
            figures_ws.add_image(image, f"D{anchor_row}")
            anchor_row += 18
        except Exception:
            continue
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


class DocxBuilder:
    """Small DOCX writer using the Office Open XML package format."""

    def __init__(self, title: str) -> None:
        self.title = title
        self.body: list[str] = []
        self.relationships: list[tuple[str, str, str]] = []
        self.media: list[tuple[str, Path]] = []
        self.next_rid = 1
        self.next_image = 1

    def add_heading(self, text: str, level: int = 1) -> None:
        """Add a Word heading paragraph."""

        style = "Title" if level == 0 else f"Heading{min(max(level, 1), 3)}"
        self.body.append(
            "<w:p>"
            f"<w:pPr><w:pStyle w:val=\"{style}\"/></w:pPr>"
            f"<w:r><w:t>{escape(text)}</w:t></w:r>"
            "</w:p>"
        )

    def add_paragraph(self, text: str, bold_prefix: str | None = None) -> None:
        """Add a plain paragraph."""

        if bold_prefix and text.startswith(bold_prefix):
            rest = text[len(bold_prefix) :]
            self.body.append(
                "<w:p>"
                f"<w:r><w:rPr><w:b/></w:rPr><w:t>{escape(bold_prefix)}</w:t></w:r>"
                f"<w:r><w:t>{escape(rest)}</w:t></w:r>"
                "</w:p>"
            )
            return
        self.body.append(f"<w:p><w:r><w:t>{escape(text)}</w:t></w:r></w:p>")

    def add_bullet(self, text: str) -> None:
        """Add a bullet-like paragraph."""

        self.body.append(f"<w:p><w:r><w:t>{escape('• ' + text)}</w:t></w:r></w:p>")

    def add_table(self, rows: list[list[Any]]) -> None:
        """Add a simple bordered table."""

        if not rows:
            return
        table = [
            "<w:tbl>",
            "<w:tblPr><w:tblStyle w:val=\"TableGrid\"/><w:tblW w:w=\"0\" w:type=\"auto\"/>"
            "<w:tblBorders>"
            "<w:top w:val=\"single\" w:sz=\"4\" w:space=\"0\" w:color=\"BFBFBF\"/>"
            "<w:left w:val=\"single\" w:sz=\"4\" w:space=\"0\" w:color=\"BFBFBF\"/>"
            "<w:bottom w:val=\"single\" w:sz=\"4\" w:space=\"0\" w:color=\"BFBFBF\"/>"
            "<w:right w:val=\"single\" w:sz=\"4\" w:space=\"0\" w:color=\"BFBFBF\"/>"
            "<w:insideH w:val=\"single\" w:sz=\"4\" w:space=\"0\" w:color=\"BFBFBF\"/>"
            "<w:insideV w:val=\"single\" w:sz=\"4\" w:space=\"0\" w:color=\"BFBFBF\"/>"
            "</w:tblBorders></w:tblPr>",
        ]
        for row_index, row in enumerate(rows):
            table.append("<w:tr>")
            for value in row:
                shading = "<w:shd w:fill=\"2B7A78\"/>" if row_index == 0 else ""
                color = "<w:color w:val=\"FFFFFF\"/>" if row_index == 0 else ""
                bold = "<w:b/>" if row_index == 0 else ""
                table.append(
                    "<w:tc>"
                    f"<w:tcPr>{shading}</w:tcPr>"
                    "<w:p><w:r>"
                    f"<w:rPr>{bold}{color}<w:sz w:val=\"18\"/></w:rPr>"
                    f"<w:t>{escape(str(value))}</w:t>"
                    "</w:r></w:p></w:tc>"
                )
            table.append("</w:tr>")
        table.append("</w:tbl>")
        self.body.append("".join(table))

    def add_image(self, path: Path, caption: str, width_inches: float = 5.9) -> None:
        """Embed an image with a caption."""

        if not path.exists():
            return
        rid = f"rId{self.next_rid}"
        self.next_rid += 1
        extension = path.suffix.lower().lstrip(".") or "png"
        media_name = f"image{self.next_image}.{extension}"
        self.next_image += 1
        self.relationships.append((rid, "http://schemas.openxmlformats.org/officeDocument/2006/relationships/image", f"media/{media_name}"))
        self.media.append((media_name, path))
        cx, cy = image_size_emu(path, width_inches)
        self.body.append(
            "<w:p><w:r><w:drawing>"
            "<wp:inline distT=\"0\" distB=\"0\" distL=\"0\" distR=\"0\" "
            "xmlns:wp=\"http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing\">"
            f"<wp:extent cx=\"{cx}\" cy=\"{cy}\"/>"
            "<wp:effectExtent l=\"0\" t=\"0\" r=\"0\" b=\"0\"/>"
            "<wp:docPr id=\"1\" name=\"Picture\"/>"
            "<wp:cNvGraphicFramePr><a:graphicFrameLocks noChangeAspect=\"1\" "
            "xmlns:a=\"http://schemas.openxmlformats.org/drawingml/2006/main\"/></wp:cNvGraphicFramePr>"
            "<a:graphic xmlns:a=\"http://schemas.openxmlformats.org/drawingml/2006/main\">"
            "<a:graphicData uri=\"http://schemas.openxmlformats.org/drawingml/2006/picture\">"
            "<pic:pic xmlns:pic=\"http://schemas.openxmlformats.org/drawingml/2006/picture\">"
            "<pic:nvPicPr><pic:cNvPr id=\"0\" name=\"Picture\"/><pic:cNvPicPr/></pic:nvPicPr>"
            "<pic:blipFill>"
            f"<a:blip r:embed=\"{rid}\" xmlns:r=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships\"/>"
            "<a:stretch><a:fillRect/></a:stretch></pic:blipFill>"
            "<pic:spPr><a:xfrm><a:off x=\"0\" y=\"0\"/>"
            f"<a:ext cx=\"{cx}\" cy=\"{cy}\"/>"
            "</a:xfrm><a:prstGeom prst=\"rect\"><a:avLst/></a:prstGeom></pic:spPr>"
            "</pic:pic></a:graphicData></a:graphic>"
            "</wp:inline></w:drawing></w:r></w:p>"
        )
        self.body.append(
            "<w:p>"
            "<w:pPr><w:jc w:val=\"center\"/></w:pPr>"
            f"<w:r><w:rPr><w:i/><w:color w:val=\"666666\"/></w:rPr><w:t>{escape(caption)}</w:t></w:r>"
            "</w:p>"
        )

    def save(self, output: Path) -> None:
        """Write the DOCX package."""

        output.parent.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as docx:
            docx.writestr("[Content_Types].xml", self.content_types_xml())
            docx.writestr("_rels/.rels", package_rels_xml())
            docx.writestr("docProps/core.xml", core_xml(self.title))
            docx.writestr("docProps/app.xml", app_xml())
            docx.writestr("word/document.xml", self.document_xml())
            docx.writestr("word/styles.xml", styles_xml())
            docx.writestr("word/_rels/document.xml.rels", self.document_rels_xml())
            for media_name, path in self.media:
                docx.write(path, f"word/media/{media_name}")

    def document_xml(self) -> str:
        """Return Word document XML."""

        section = (
            "<w:sectPr>"
            "<w:pgSz w:w=\"11906\" w:h=\"16838\"/>"
            "<w:pgMar w:top=\"1134\" w:right=\"992\" w:bottom=\"1134\" w:left=\"992\" "
            "w:header=\"708\" w:footer=\"708\" w:gutter=\"0\"/>"
            "</w:sectPr>"
        )
        return (
            "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
            "<w:document xmlns:w=\"http://schemas.openxmlformats.org/wordprocessingml/2006/main\" "
            "xmlns:r=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships\">"
            f"<w:body>{''.join(self.body)}{section}</w:body></w:document>"
        )

    def document_rels_xml(self) -> str:
        """Return document relationships XML."""

        entries = [
            '<Relationship Id="rIdStyles" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
        ]
        for rid, rel_type, target in self.relationships:
            entries.append(f'<Relationship Id="{rid}" Type="{rel_type}" Target="{escape(target)}"/>')
        return (
            "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
            "<Relationships xmlns=\"http://schemas.openxmlformats.org/package/2006/relationships\">"
            f"{''.join(entries)}</Relationships>"
        )

    def content_types_xml(self) -> str:
        """Return content types XML."""

        media_defaults = set(path.suffix.lower().lstrip(".") or "png" for _, path in self.media)
        defaults = [
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>',
            '<Default Extension="xml" ContentType="application/xml"/>',
        ]
        content_map = {
            "png": "image/png",
            "jpg": "image/jpeg",
            "jpeg": "image/jpeg",
            "gif": "image/gif",
            "bmp": "image/bmp",
        }
        for extension in sorted(media_defaults | {"png"}):
            defaults.append(f'<Default Extension="{extension}" ContentType="{content_map.get(extension, "application/octet-stream")}"/>')
        overrides = [
            '<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>',
            '<Override PartName="/word/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.styles+xml"/>',
            '<Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>',
            '<Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>',
        ]
        return (
            "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
            "<Types xmlns=\"http://schemas.openxmlformats.org/package/2006/content-types\">"
            f"{''.join(defaults)}{''.join(overrides)}</Types>"
        )


def image_size_emu(path: Path, width_inches: float) -> tuple[int, int]:
    """Return image size in EMUs preserving aspect ratio."""

    width_px = 1600
    height_px = 900
    try:
        from PIL import Image

        with Image.open(path) as image:
            width_px, height_px = image.size
    except Exception:
        pass
    cx = int(width_inches * 914400)
    cy = int(cx * height_px / max(width_px, 1))
    return cx, cy


def package_rels_xml() -> str:
    """Return root package relationships."""

    return (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<Relationships xmlns=\"http://schemas.openxmlformats.org/package/2006/relationships\">"
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>'
        '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>'
        "</Relationships>"
    )


def core_xml(title: str) -> str:
    """Return core document properties."""

    now = datetime.now(UTC).strftime("%Y-%m-%dT%H:%M:%SZ")
    return (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        '<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
        'xmlns:dc="http://purl.org/dc/elements/1.1/" '
        'xmlns:dcterms="http://purl.org/dc/terms/" '
        'xmlns:dcmitype="http://purl.org/dc/dcmitype/" '
        'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
        f"<dc:title>{escape(title)}</dc:title><dc:creator>report_generator.py</dc:creator>"
        f'<dcterms:created xsi:type="dcterms:W3CDTF">{now}</dcterms:created>'
        f'<dcterms:modified xsi:type="dcterms:W3CDTF">{now}</dcterms:modified>'
        "</cp:coreProperties>"
    )


def app_xml() -> str:
    """Return app document properties."""

    return (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" '
        'xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">'
        "<Application>Python</Application></Properties>"
    )


def styles_xml() -> str:
    """Return a compact Word styles document."""

    return (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        '<w:styles xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        '<w:style w:type="paragraph" w:default="1" w:styleId="Normal">'
        '<w:name w:val="Normal"/><w:qFormat/><w:rPr><w:rFonts w:ascii="Microsoft YaHei" w:eastAsia="Microsoft YaHei"/>'
        '<w:sz w:val="21"/></w:rPr></w:style>'
        '<w:style w:type="paragraph" w:styleId="Title"><w:name w:val="Title"/><w:basedOn w:val="Normal"/>'
        '<w:qFormat/><w:pPr><w:jc w:val="center"/></w:pPr><w:rPr><w:b/><w:sz w:val="34"/></w:rPr></w:style>'
        '<w:style w:type="paragraph" w:styleId="Heading1"><w:name w:val="heading 1"/><w:basedOn w:val="Normal"/>'
        '<w:qFormat/><w:pPr><w:spacing w:before="320" w:after="120"/></w:pPr><w:rPr><w:b/><w:sz w:val="28"/></w:rPr></w:style>'
        '<w:style w:type="paragraph" w:styleId="Heading2"><w:name w:val="heading 2"/><w:basedOn w:val="Normal"/>'
        '<w:qFormat/><w:pPr><w:spacing w:before="240" w:after="80"/></w:pPr><w:rPr><w:b/><w:sz w:val="24"/></w:rPr></w:style>'
        '<w:style w:type="table" w:styleId="TableGrid"><w:name w:val="Table Grid"/><w:basedOn w:val="TableNormal"/>'
        '<w:uiPriority w:val="59"/><w:rsid w:val="00000000"/><w:tblPr><w:tblBorders>'
        '<w:top w:val="single" w:sz="4" w:space="0" w:color="auto"/>'
        '<w:left w:val="single" w:sz="4" w:space="0" w:color="auto"/>'
        '<w:bottom w:val="single" w:sz="4" w:space="0" w:color="auto"/>'
        '<w:right w:val="single" w:sz="4" w:space="0" w:color="auto"/>'
        '<w:insideH w:val="single" w:sz="4" w:space="0" w:color="auto"/>'
        '<w:insideV w:val="single" w:sz="4" w:space="0" w:color="auto"/>'
        "</w:tblBorders></w:tblPr></w:style>"
        "</w:styles>"
    )


def report_text_summary(
    args: argparse.Namespace,
    source: SourceInfo,
    training_rows: list[dict[str, str]],
    ablation_rows: list[dict[str, Any]],
    train_args: dict[str, Any],
    data_info: dict[str, Any],
) -> dict[str, str]:
    """Build reusable report text fragments."""

    epochs = train_args.get("epochs") or (training_rows[-1].get("epoch") if training_rows else "")
    imgsz = train_args.get("imgsz", "")
    batch = train_args.get("batch", "")
    device = train_args.get("device", "")
    model = train_args.get("model", "")
    best = best_epoch(training_rows)
    best_accuracy = best_row(ablation_rows, "mAP50-95")
    best_speed = best_row(ablation_rows, "FPS")
    deployed = next((row for row in ablation_rows if row.get("Experiment") == "Baseline+CBAM+BiFPN"), None)
    if deployed is None:
        deployed = best_accuracy
    focal_rows = [row for row in ablation_rows if truthy(row.get("FocalLoss"))]
    focal_negative = bool(
        focal_rows
        and best_accuracy
        and all((numeric(row.get("mAP50-95")) or 0.0) < (numeric(best_accuracy.get("mAP50-95")) or 0.0) for row in focal_rows)
    )
    return {
        "source_line": f"训练曲线读取自 {source.results_csv.relative_to(ROOT)}；消融结果读取自 {source.ablation_summary.relative_to(ROOT) if source.ablation_summary else '未发现 summary.csv'}。",
        "setup": (
            f"本次主训练运行使用模型 {model or 'YOLOv8'}，训练轮数 {epochs or '-'}，输入尺寸 {imgsz or '-'}，"
            f"batch={batch or '-'}，device={device or '-'}。"
        ),
        "dataset": (
            f"数据集配置文件为 {data_info.get('dataset_yaml') or '未发现'}，共识别 {data_info.get('classes', 0)} 个类别。"
            f"当前可统计图像数为训练集 {data_info.get('train_images', 0)} 张、验证集 {data_info.get('val_images', 0)} 张。"
        ),
        "training": (
            f"主训练在第 {best.get('best_epoch') or '-'} 轮取得最佳 mAP50-95={fmt(best.get('best_mAP50-95'))}，"
            f"对应 mAP50={fmt(best.get('best_mAP50'))}。最终一轮 box_loss={fmt(best.get('final_box_loss'))}，"
            f"cls_loss={fmt(best.get('final_cls_loss'))}，dfl_loss={fmt(best.get('final_dfl_loss'))}。"
        ),
        "accuracy": (
            f"消融实验中，精度最优模型为 {best_accuracy.get('Experiment') if best_accuracy else '-'}，"
            f"mAP50-95={fmt(best_accuracy.get('mAP50-95') if best_accuracy else None)}，"
            f"mAP50={fmt(best_accuracy.get('mAP50') if best_accuracy else None)}。"
        ),
        "speed": (
            f"速度最优模型为 {best_speed.get('Experiment') if best_speed else '-'}，"
            f"FPS={fmt(best_speed.get('FPS') if best_speed else None, 2)}。"
        ),
        "choice": (
            f"综合精度与速度，推荐默认部署模型为 {deployed.get('Experiment') if deployed else '-'}，"
            f"其 mAP50-95={fmt(deployed.get('mAP50-95') if deployed else None)}，"
            f"FPS={fmt(deployed.get('FPS') if deployed else None, 2)}。"
        ),
        "focal": (
            "Focal Loss 相关组合在当前结果中未成为最优精度模型，可作为负向消融结论保留。"
            if focal_negative
            else "Focal Loss 相关组合的影响需要结合完整表格继续比较。"
        ),
        "class_names": str(data_info.get("class_names", "")),
    }


def docx_table_rows(ablation_rows: list[dict[str, Any]]) -> list[list[str]]:
    """Build compact ablation rows for the Word report."""

    rows = [["Rank", "Experiment", "Precision", "Recall", "mAP50", "mAP50-95", "FPS", "ΔmAP50-95"]]
    for row in ablation_rows[:12]:
        rows.append(
            [
                str(row.get("Rank", "")),
                str(row.get("Experiment", "")),
                fmt(row.get("Precision")),
                fmt(row.get("Recall")),
                fmt(row.get("mAP50")),
                fmt(row.get("mAP50-95")),
                fmt(row.get("FPS"), 2),
                fmt_delta(row.get("Delta_mAP50-95")),
            ]
        )
    return rows


def write_report_docx(
    output: Path,
    args: argparse.Namespace,
    source: SourceInfo,
    training_rows: list[dict[str, str]],
    ablation_rows: list[dict[str, Any]],
    train_args: dict[str, Any],
    data_info: dict[str, Any],
    figures: FigureSet,
) -> None:
    """Write the final Word report."""

    text = report_text_summary(args, source, training_rows, ablation_rows, train_args, data_info)
    doc = DocxBuilder(args.report_title)
    doc.add_heading(args.report_title, level=0)
    doc.add_paragraph(args.project_name)
    doc.add_paragraph(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    doc.add_paragraph(text["source_line"])

    doc.add_heading("一、项目与数据概况", level=1)
    doc.add_paragraph(
        "本项目面向中医药饮片图像检测与识别任务，采用 YOLOv8n 作为基线模型，并围绕注意力机制、"
        "多尺度特征融合、轻量化卷积、解耦检测头和分类损失函数进行结构消融。报告自动汇总训练日志、"
        "验证指标、混淆矩阵与消融实验结果，用于支撑模型选型和实验材料归档。"
    )
    doc.add_paragraph(text["dataset"])
    if text["class_names"]:
        doc.add_paragraph(f"类别集合：{text['class_names']}")

    doc.add_heading("二、实验设置", level=1)
    doc.add_paragraph(text["setup"])
    doc.add_bullet("训练输出固定读取 YOLO results.csv，包含 box_loss、cls_loss、dfl_loss、Precision、Recall、mAP50 和 mAP50-95 等逐轮指标。")
    doc.add_bullet("消融实验表来自 summary.csv；若存在 history.csv，则同步绘制多实验收敛曲线。")
    doc.add_bullet("所有生成材料均保存为 CSV、Excel、PNG 与 Word 格式，便于论文、答辩和项目验收复用。")

    doc.add_heading("三、训练收敛分析", level=1)
    doc.add_paragraph(text["training"])
    doc.add_paragraph(
        "Loss 曲线用于观察定位框回归、类别分类和分布式焦点损失的收敛稳定性。若训练集和验证集曲线同步下降且后期趋于平稳，"
        "说明模型已较好地学习饮片形态、纹理和边界特征；若验证损失回升，则需要进一步检查过拟合或数据分布差异。"
    )
    doc.add_image(figures.loss.get("box_loss", Path()), "图 1  box_loss 曲线")
    doc.add_image(figures.loss.get("cls_loss", Path()), "图 2  cls_loss 曲线")
    doc.add_image(figures.loss.get("dfl_loss", Path()), "图 3  dfl_loss 曲线")

    doc.add_heading("四、检测精度分析", level=1)
    doc.add_paragraph(
        "mAP50 反映 IoU=0.50 阈值下的检测能力，mAP50-95 则在更严格的多阈值范围内评价定位质量，"
        "因此后者更适合作为最终模型排序和部署选型的核心指标。"
    )
    doc.add_image(figures.map.get("map50", Path()), "图 4  mAP50 曲线")
    doc.add_image(figures.map.get("map50_95", Path()), "图 5  mAP50-95 曲线")
    doc.add_paragraph(text["accuracy"])

    doc.add_heading("五、混淆矩阵分析", level=1)
    doc.add_paragraph(
        "混淆矩阵用于观察各类别之间的误检和漏检关系。对于外观相似、颜色和切片形态接近的饮片类别，"
        "该图可以帮助定位后续数据增强、类别重采样和采集规范优化的重点。"
    )
    doc.add_image(figures.confusion.get("confusion_matrix.png", Path()), "图 6  混淆矩阵")
    if "confusion_matrix_normalized.png" in figures.confusion:
        doc.add_image(figures.confusion["confusion_matrix_normalized.png"], "图 7  归一化混淆矩阵")

    doc.add_heading("六、消融实验结果", level=1)
    doc.add_paragraph(
        "消融实验从 Baseline 出发，分别考察 CBAM、BiFPN、Focal Loss、GhostConv 与 Decoupled Head 对精度和速度的影响。"
        "表中 ΔmAP50-95 以 Baseline 为参照，正值表示相对提升。"
    )
    if ablation_rows:
        doc.add_table(docx_table_rows(ablation_rows))
    else:
        doc.add_paragraph("未发现可用的消融实验 summary.csv。")
    doc.add_image(figures.ablation.get("ablation_map50_95_bar", Path()), "图 8  消融实验 mAP50-95 排名")
    doc.add_image(figures.ablation.get("ablation_speed_accuracy", Path()), "图 9  精度-速度权衡")
    doc.add_paragraph(text["speed"])
    doc.add_paragraph(text["focal"])

    doc.add_heading("七、结论与模型选型", level=1)
    doc.add_paragraph(text["choice"])
    doc.add_paragraph(
        "从当前结果看，注意力机制与特征融合结构能提升模型对局部纹理和多尺度目标的表达能力；轻量化结构可作为速度与模型复杂度的对照；"
        "复杂组合并不必然带来精度提升，尤其在样本规模有限的情况下，需要通过公平预训练迁移、统一训练轮数和固定验证集进行验证。"
    )
    doc.add_paragraph(
        "后续建议围绕三点继续完善：一是扩充易混类别样本并进行针对性增强；二是对最终部署模型导出 ONNX/OpenVINO/NCNN 后实测端侧速度；"
        "三是保留完整训练日志和混淆矩阵，形成可追溯的模型迭代档案。"
    )
    doc.save(output)


def source_rows(source: SourceInfo) -> list[dict[str, str]]:
    """Return source information rows for exports."""

    rows = [
        {"Item": "Primary run", "Path": str(source.run_dir)},
        {"Item": "Training results", "Path": str(source.results_csv)},
    ]
    if source.args_yaml:
        rows.append({"Item": "Training args", "Path": str(source.args_yaml)})
    if source.ablation_dir:
        rows.append({"Item": "Ablation directory", "Path": str(source.ablation_dir)})
    if source.ablation_summary:
        rows.append({"Item": "Ablation summary", "Path": str(source.ablation_summary)})
    if source.ablation_history:
        rows.append({"Item": "Ablation history", "Path": str(source.ablation_history)})
    return rows


def collect_figure_paths(figures: FigureSet) -> list[Path]:
    """Flatten figure paths in a stable order."""

    paths: list[Path] = []
    for group in [figures.loss, figures.map, figures.confusion, figures.ablation]:
        for path in group.values():
            if path.exists() and path not in paths:
                paths.append(path)
    return paths


def main() -> int:
    """Generate all experiment materials."""

    args = parse_args()
    output_dir = resolve_path(args.output)
    assert output_dir is not None
    data_dir = output_dir / "csv"
    figures_dir = output_dir / "png"
    excel_dir = output_dir / "excel"
    report_dir = output_dir / "word"
    source = resolve_sources(args)

    training_rows = read_csv_rows(source.results_csv)
    if not training_rows:
        raise RuntimeError(f"No training rows found in {source.results_csv}")
    ablation_raw_rows = merge_final_selection_rows(read_csv_rows(source.ablation_summary))
    history_rows = read_csv_rows(source.ablation_history)
    ablation_rows = enrich_ablation_rows(ablation_raw_rows)
    train_args = read_simple_yaml(source.args_yaml)
    data_info = dataset_summary()

    write_csv_rows(data_dir / "training_metrics.csv", training_rows)
    write_csv_rows(data_dir / "ablation_summary.csv", ablation_rows, ABLATION_FIELDS)
    if history_rows:
        write_csv_rows(data_dir / "ablation_history.csv", history_rows)
    write_csv_rows(data_dir / "sources.csv", source_rows(source), ["Item", "Path"])

    figures = generate_figures(training_rows, history_rows, ablation_rows, source, figures_dir)
    figure_paths = collect_figure_paths(figures)
    excel_path = excel_dir / "experiment_materials.xlsx"
    export_excel(excel_path, training_rows, ablation_rows, history_rows, source_rows(source), figure_paths)

    docx_path = report_dir / "experiment_report.docx"
    write_report_docx(docx_path, args, source, training_rows, ablation_rows, train_args, data_info, figures)

    print("Experiment materials generated successfully.")
    print(f"Output directory: {output_dir}")
    print(f"CSV: {data_dir}")
    print(f"Excel: {excel_path}")
    print(f"PNG: {figures_dir}")
    print(f"Word: {docx_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


