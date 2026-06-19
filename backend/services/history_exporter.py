"""Excel export helpers for detection history."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ..config import BACKEND_ROOT
from ..models import DetectionRecord
from .history import load_json


EXPORT_DIR = BACKEND_ROOT / "data" / "exports"


def export_history_to_excel(records: list[DetectionRecord]) -> Path:
    """Write history records to an Excel file and return its path."""

    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    path = EXPORT_DIR / f"history_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid4().hex[:8]}.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Detection History"

    headers = [
        "ID",
        "Created At",
        "Mode",
        "Source Type",
        "Source Path",
        "Output Path",
        "Model Path",
        "Device",
        "Elapsed(ms)",
        "FPS",
        "Count",
        "Success",
        "Status",
        "Class Counts",
        "Error",
    ]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for record in records:
        class_counts = load_json(record.class_counts_json, {})
        class_text = ", ".join(f"{name}: {count}" for name, count in class_counts.items())
        sheet.append(
            [
                record.id,
                record.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                record.mode,
                record.source_type,
                record.source_path,
                record.output_path or "",
                record.model_path,
                record.device,
                round(record.elapsed_ms, 2),
                round(record.fps, 2),
                record.total_count,
                "yes" if record.success else "no",
                record.status,
                class_text,
                record.error_message or "",
            ]
        )

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        width = min(max(max_length + 2, 10), 60)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    sheet.freeze_panes = "A2"
    workbook.save(path)
    return path
